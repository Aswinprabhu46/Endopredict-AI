#!/usr/bin/env python3
"""
===============================================================================
EndoPredict AI — Concurrent Load Test Suite (100 Virtual Users × 1 Minute)
===============================================================================
Simulates 100 concurrent virtual users sending continuous HTTP requests to the
EndoPredict web application for 60 seconds.  Measures:
  • Requests per Second (RPS)
  • Response-time distribution (min / avg / p50 / p95 / p99 / max)
  • HTTP status-code breakdown
  • Per-endpoint statistics
  • Error rate

Generates an Excel report: EndoPredict_Load_Test_Report.xlsx
===============================================================================
"""

import asyncio
import aiohttp
import time
import statistics
import os
import sys
import traceback
from datetime import datetime
from collections import defaultdict

# openpyxl for Excel report
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_URL       = os.getenv("LOAD_TEST_URL", "http://localhost:5173")
VIRTUAL_USERS  = int(os.getenv("VIRTUAL_USERS", "100"))
DURATION_SECS  = int(os.getenv("DURATION_SECS", "60"))
REPORT_NAME    = "EndoPredict_Load_Test_Report.xlsx"

# Endpoints to test (relative paths)
ENDPOINTS = [
    "/",
    "/index.html",
    "/manifest.json",
    "/favicon.svg",
    "/icons.svg",
    "/sw.js",
]


# ── Collector ─────────────────────────────────────────────────────────────────
class LoadTestCollector:
    """Thread-safe (asyncio-safe) collector for request results."""

    def __init__(self):
        self.results = []                          # list of dicts
        self.status_counts = defaultdict(int)       # {status_code: count}
        self.endpoint_results = defaultdict(list)   # {endpoint: [latency, …]}
        self.errors = 0
        self.total = 0
        self._lock = asyncio.Lock()

    async def record(self, endpoint, status, latency_ms, error_msg=""):
        async with self._lock:
            self.total += 1
            self.status_counts[status] += 1
            if status == 0 or status >= 400:
                self.errors += 1
            self.endpoint_results[endpoint].append(latency_ms)
            self.results.append({
                "id": self.total,
                "endpoint": endpoint,
                "status": status,
                "latency_ms": round(latency_ms, 2),
                "error": error_msg,
                "timestamp": datetime.now().strftime("%H:%M:%S.%f")[:-3],
            })


# ── Single virtual-user coroutine ─────────────────────────────────────────────
async def virtual_user(user_id: int, session: aiohttp.ClientSession,
                       collector: LoadTestCollector, stop_event: asyncio.Event):
    """
    One virtual user that loops through ENDPOINTS repeatedly until stop_event
    is set (after DURATION_SECS seconds).
    """
    idx = 0
    while not stop_event.is_set():
        endpoint = ENDPOINTS[idx % len(ENDPOINTS)]
        url = f"{BASE_URL}{endpoint}"
        t0 = time.perf_counter()
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                await resp.read()
                latency = (time.perf_counter() - t0) * 1000
                await collector.record(endpoint, resp.status, latency)
        except asyncio.CancelledError:
            break
        except Exception as exc:
            latency = (time.perf_counter() - t0) * 1000
            await collector.record(endpoint, 0, latency, str(exc)[:120])
        idx += 1
        await asyncio.sleep(0)  # yield to event loop


# ── Orchestrator ──────────────────────────────────────────────────────────────
async def run_load_test():
    print("\n" + "=" * 80)
    print("🚀 ENDOPREDICT CONCURRENT LOAD TEST")
    print("=" * 80)
    print(f"  Target URL      : {BASE_URL}")
    print(f"  Virtual Users   : {VIRTUAL_USERS}")
    print(f"  Duration        : {DURATION_SECS}s")
    print(f"  Endpoints       : {len(ENDPOINTS)}")
    print(f"  Started At      : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80 + "\n")

    collector = LoadTestCollector()
    stop_event = asyncio.Event()

    connector = aiohttp.TCPConnector(limit=0, limit_per_host=0)
    async with aiohttp.ClientSession(connector=connector) as session:
        # Warm-up request
        try:
            async with session.get(f"{BASE_URL}/") as r:
                await r.read()
                print(f"  ✅ Warm-up request: HTTP {r.status}\n")
        except Exception as e:
            print(f"  ⚠️  Warm-up failed: {e}\n")

        # Spawn virtual users
        tasks = []
        for uid in range(VIRTUAL_USERS):
            tasks.append(asyncio.create_task(
                virtual_user(uid, session, collector, stop_event)
            ))

        # Progress reporting every 10 seconds
        start = time.perf_counter()
        for tick in range(DURATION_SECS // 10):
            await asyncio.sleep(10)
            elapsed = time.perf_counter() - start
            rps = collector.total / elapsed if elapsed > 0 else 0
            print(f"  ⏱  {int(elapsed)}s elapsed  |  {collector.total} requests  |  "
                  f"{rps:.0f} req/s  |  errors: {collector.errors}")

        remaining = DURATION_SECS - (DURATION_SECS // 10) * 10
        if remaining > 0:
            await asyncio.sleep(remaining)

        # Signal stop
        stop_event.set()
        # Give tasks a moment to finish gracefully
        await asyncio.sleep(0.5)
        for t in tasks:
            t.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)

    total_elapsed = time.perf_counter() - start
    return collector, total_elapsed


# ── Statistics helpers ────────────────────────────────────────────────────────
def percentile(data, p):
    if not data:
        return 0
    k = (len(data) - 1) * (p / 100)
    f = int(k)
    c = f + 1
    if c >= len(data):
        return data[-1]
    return data[f] + (data[c] - data[f]) * (k - f)


def compute_stats(latencies):
    if not latencies:
        return {"min": 0, "max": 0, "avg": 0, "p50": 0, "p95": 0, "p99": 0, "count": 0}
    s = sorted(latencies)
    return {
        "count": len(s),
        "min": round(s[0], 2),
        "max": round(s[-1], 2),
        "avg": round(statistics.mean(s), 2),
        "p50": round(percentile(s, 50), 2),
        "p95": round(percentile(s, 95), 2),
        "p99": round(percentile(s, 99), 2),
    }


# ── Excel report generation ──────────────────────────────────────────────────
def generate_report(collector: LoadTestCollector, elapsed: float):
    wb = openpyxl.Workbook()

    # Styles
    title_font    = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font     = Font(name="Calibri", size=11, bold=True)
    regular_font  = Font(name="Calibri", size=10)
    navy_fill     = PatternFill("solid", fgColor="1B365D")
    blue_fill     = PatternFill("solid", fgColor="1A73E8")
    green_fill    = PatternFill("solid", fgColor="D4EDDA")
    red_fill      = PatternFill("solid", fgColor="F8D7DA")
    grey_fill     = PatternFill("solid", fgColor="F1F3F4")
    pass_font     = Font(name="Calibri", size=10, bold=True, color="155724")
    fail_font     = Font(name="Calibri", size=10, bold=True, color="721C24")
    thin_border   = Border(
        left=Side("thin", "D0D0D0"), right=Side("thin", "D0D0D0"),
        top=Side("thin", "D0D0D0"), bottom=Side("thin", "D0D0D0"),
    )
    center = Alignment(horizontal="center", vertical="center")

    all_latencies = [r["latency_ms"] for r in collector.results]
    overall = compute_stats(all_latencies)
    rps = collector.total / elapsed if elapsed > 0 else 0
    error_rate = (collector.errors / collector.total * 100) if collector.total else 0

    # ── Sheet 1 — Executive Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws.merge_cells("A1:G2")
    c = ws["A1"]
    c.value = f"  🦷 EndoPredict Load Test Report — {VIRTUAL_USERS} Users × {DURATION_SECS}s"
    c.font = title_font; c.fill = navy_fill; c.alignment = Alignment(vertical="center")

    # KPI cards
    kpis = [
        ("Target URL",       BASE_URL,                     "B4", "C4"),
        ("Virtual Users",    VIRTUAL_USERS,                "E4", "F4"),
        ("Duration (s)",     f"{elapsed:.1f}",             "B6", "C6"),
        ("Total Requests",   collector.total,              "E6", "F6"),
        ("Requests / sec",   f"{rps:.1f}",                "B8", "C8"),
        ("Error Rate",       f"{error_rate:.2f}%",        "E8", "F8"),
        ("Avg Latency (ms)", f"{overall['avg']:.1f}",     "B10", "C10"),
        ("Min Latency (ms)", f"{overall['min']:.1f}",     "E10", "F10"),
        ("P50 Latency (ms)", f"{overall['p50']:.1f}",     "B12", "C12"),
        ("P95 Latency (ms)", f"{overall['p95']:.1f}",     "E12", "F12"),
        ("P99 Latency (ms)", f"{overall['p99']:.1f}",     "B14", "C14"),
        ("Max Latency (ms)", f"{overall['max']:.1f}",     "E14", "F14"),
    ]
    for label, val, lc, vc in kpis:
        ws[lc].value = label; ws[lc].font = bold_font; ws[lc].fill = grey_fill
        ws[lc].alignment = center; ws[lc].border = thin_border
        ws[vc].value = val; ws[vc].font = bold_font
        ws[vc].alignment = center; ws[vc].border = thin_border

    # HTTP status breakdown
    row = 17
    ws.cell(row, 1, "HTTP Status Code Breakdown").font = Font(size=13, bold=True, color="1B365D")
    row += 1
    for ci, h in enumerate(["Status Code", "Count", "Percentage"], 1):
        c = ws.cell(row, ci, h); c.font = header_font; c.fill = blue_fill
        c.alignment = center; c.border = thin_border
    row += 1
    for code in sorted(collector.status_counts.keys()):
        cnt = collector.status_counts[code]
        pct = cnt / collector.total * 100 if collector.total else 0
        for ci, v in enumerate([code if code else "Error/Timeout", cnt, f"{pct:.1f}%"], 1):
            c = ws.cell(row, ci, v); c.font = regular_font
            c.alignment = center; c.border = thin_border
        row += 1

    # Per-endpoint stats
    row += 1
    ws.cell(row, 1, "Per-Endpoint Statistics").font = Font(size=13, bold=True, color="1B365D")
    row += 1
    ep_headers = ["Endpoint", "Requests", "Avg (ms)", "Min (ms)", "P50 (ms)",
                  "P95 (ms)", "P99 (ms)", "Max (ms)"]
    for ci, h in enumerate(ep_headers, 1):
        c = ws.cell(row, ci, h); c.font = header_font; c.fill = blue_fill
        c.alignment = center; c.border = thin_border
    row += 1
    for ep in sorted(collector.endpoint_results.keys()):
        st = compute_stats(collector.endpoint_results[ep])
        vals = [ep, st["count"], st["avg"], st["min"], st["p50"],
                st["p95"], st["p99"], st["max"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci, v); c.font = regular_font
            c.alignment = center if ci > 1 else Alignment(horizontal="left")
            c.border = thin_border
        row += 1

    # Auto-width
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mx + 4, 14)

    # ── Sheet 2 — Detailed Request Log (sample first 500 rows) ───────────────
    ws2 = wb.create_sheet(f"Request Log ({min(len(collector.results), 500)} rows)")
    headers = ["#", "Endpoint", "HTTP Status", "Latency (ms)", "Result", "Timestamp", "Error"]
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(1, ci, h); c.font = header_font; c.fill = navy_fill
        c.alignment = center; c.border = thin_border

    for ri, r in enumerate(collector.results[:500], 2):
        is_ok = r["status"] == 200
        vals = [r["id"], r["endpoint"], r["status"], r["latency_ms"],
                "PASS" if is_ok else "FAIL", r["timestamp"], r["error"]]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci, v); c.font = regular_font; c.border = thin_border
            if ci == 5:
                c.fill = green_fill if is_ok else red_fill
                c.font = pass_font if is_ok else fail_font
            c.alignment = center if ci != 7 else Alignment(horizontal="left")

    for col in ws2.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 4, 12), 50)

    path = os.path.join(os.getcwd(), REPORT_NAME)
    wb.save(path)
    return path


# ── Print summary ─────────────────────────────────────────────────────────────
def print_summary(collector, elapsed):
    all_lat = sorted(r["latency_ms"] for r in collector.results)
    rps = collector.total / elapsed if elapsed else 0
    err_rate = collector.errors / collector.total * 100 if collector.total else 0

    print("\n" + "=" * 80)
    print("📊 LOAD TEST RESULTS")
    print("=" * 80)
    print(f"  Total Requests   : {collector.total}")
    print(f"  Requests / sec   : {rps:.1f} req/s")
    print(f"  Error Rate       : {err_rate:.2f}%")
    print(f"  Duration         : {elapsed:.1f}s")
    print()
    s = compute_stats(all_lat)
    print(f"  Response Time:")
    print(f"    Min            : {s['min']:.1f} ms")
    print(f"    Average        : {s['avg']:.1f} ms")
    print(f"    P50 (Median)   : {s['p50']:.1f} ms")
    print(f"    P95            : {s['p95']:.1f} ms")
    print(f"    P99            : {s['p99']:.1f} ms")
    print(f"    Max            : {s['max']:.1f} ms")
    print()
    print("  HTTP Status Codes:")
    for code in sorted(collector.status_counts.keys()):
        print(f"    {code if code else 'Error'}: {collector.status_counts[code]}")
    print("=" * 80)


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        collector, elapsed = asyncio.run(run_load_test())
        print_summary(collector, elapsed)
        path = generate_report(collector, elapsed)
        print(f"\n📁 Excel Report: {path}")
        print("✅ Load test completed successfully!")
    except Exception as e:
        print(f"\n❌ Load test failed: {e}")
        traceback.print_exc()
        sys.exit(1)
