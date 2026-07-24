#!/usr/bin/env python3
"""
===============================================================================
EndoPredict AI — Static Security Code Review (SAST)
===============================================================================
Scans the source code for insecure coding patterns, misconfigurations, and
common vulnerability classes. Produces:
  • Console summary
  • Excel report: EndoPredict_Security_Review_Report.xlsx
  • Markdown executive summary in Vulnerability Test Results/

This is a STATIC code analysis tool — it reads source files and matches
patterns. It does NOT send network traffic or exploit anything.
===============================================================================
"""

import os
import re
import json
import sys
import traceback
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DIR      = os.path.join(PROJECT_ROOT, "src")
PUBLIC_DIR   = os.path.join(PROJECT_ROOT, "public")
REPORT_NAME  = "EndoPredict_Security_Review_Report.xlsx"
SUMMARY_DIR  = os.path.join(PROJECT_ROOT, "Vulnerability Test Results")
EXTENSIONS   = {".js", ".jsx", ".ts", ".tsx", ".json", ".html", ".env"}


# ── Finding dataclass ─────────────────────────────────────────────────────────
class Finding:
    def __init__(self, fid, severity, category, vuln_type, file_path,
                 line_no, code_snippet, explanation, remediation):
        self.fid          = fid
        self.severity     = severity      # Critical / High / Medium / Low
        self.category     = category
        self.vuln_type    = vuln_type
        self.file_path    = file_path
        self.line_no      = line_no
        self.code_snippet = code_snippet[:200]
        self.explanation  = explanation
        self.remediation  = remediation


# ── Scanner ───────────────────────────────────────────────────────────────────
class SecurityScanner:
    def __init__(self, project_root):
        self.root     = project_root
        self.findings = []
        self.counter  = 0

    def _next_id(self):
        self.counter += 1
        return f"SEC-{self.counter:03d}"

    def add(self, severity, category, vuln_type, fpath, line_no,
            snippet, explanation, remediation):
        self.findings.append(Finding(
            self._next_id(), severity, category, vuln_type,
            os.path.relpath(fpath, self.root), line_no, snippet.strip(),
            explanation, remediation
        ))

    def scan_file(self, fpath):
        """Run all pattern checks against a single file."""
        try:
            with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
                lines   = content.split("\n")
        except Exception:
            return

        rel = os.path.relpath(fpath, self.root)

        for i, line in enumerate(lines, 1):
            stripped = line.strip()

            # ── 1. Authentication ────────────────────────────────────────
            # Plaintext password storage
            if re.search(r'password\s*[:=]\s*["\'](?!•)', stripped, re.I):
                if "type=" not in stripped and "placeholder" not in stripped:
                    self.add("Critical", "Authentication",
                             "Plaintext Password Storage", fpath, i, stripped,
                             "Passwords stored/compared in plaintext without hashing",
                             "Hash passwords with bcrypt/argon2 before storage")

            # Plaintext password comparison
            if re.search(r'\.password\s*===?\s*password', stripped):
                self.add("Critical", "Authentication",
                         "Plaintext Password Comparison", fpath, i, stripped,
                         "Direct string comparison of passwords instead of hash verification",
                         "Use bcrypt.compare() or similar constant-time hash comparison")

            # No session expiry / no token TTL
            if "CURRENT_USER" in stripped and "setItem" in stripped:
                self.add("High", "Authentication",
                         "No Session Expiry", fpath, i, stripped,
                         "User session stored in localStorage without expiry timestamp",
                         "Add expiry timestamp and validate on each load; use sessionStorage or HttpOnly cookies")

            # Demo credentials in source code
            if re.search(r'password123', stripped):
                self.add("High", "Authentication",
                         "Hardcoded Demo Credentials", fpath, i, stripped,
                         "Default weak password 'password123' hardcoded in source",
                         "Remove hardcoded credentials; use env vars or first-run setup flow")

            # ── 2. Authorization ─────────────────────────────────────────
            # Client-side only auth check
            if "getCurrentUser" in stripped and "localStorage" in stripped:
                self.add("High", "Authorization",
                         "Client-Side Only Auth", fpath, i, stripped,
                         "Authentication state relies solely on client-side localStorage",
                         "Implement server-side session validation with Firebase Auth tokens")

            # No role-based access control
            if "registerUser" in stripped and "role" not in content[:5000]:
                self.add("Medium", "Authorization",
                         "Missing Role-Based Access Control", fpath, i, stripped,
                         "User registration has no role field; all users have same privileges",
                         "Add role field (admin/doctor/nurse) and enforce role checks")

            # ── 3. Sensitive Data Exposure ────────────────────────────────
            # Firebase config values hardcoded
            if re.search(r'(messagingSenderId|appId)\s*:\s*"[0-9]', stripped):
                self.add("Medium", "Sensitive Data Exposure",
                         "Hardcoded Firebase Identifiers", fpath, i, stripped,
                         "Firebase project identifiers committed to source control",
                         "Move to .env file; add .env to .gitignore; use VITE_* env vars")

            # API key in URL query parameter
            if re.search(r'key=\$\{.*apiKey', stripped):
                self.add("High", "Sensitive Data Exposure",
                         "API Key Exposed in URL Query String", fpath, i, stripped,
                         "Gemini API key passed as URL query parameter, visible in browser history/logs",
                         "Proxy API calls through a backend server; never expose API keys client-side")

            # API key stored in localStorage
            if "endopredict_custom_gemini_key" in stripped and "setItem" in stripped:
                self.add("Medium", "Sensitive Data Exposure",
                         "API Key in localStorage", fpath, i, stripped,
                         "User-supplied API key stored in localStorage without encryption",
                         "Encrypt before storage or proxy through authenticated backend")

            # Sensitive data in console.log
            if re.search(r'console\.(log|error|warn).*(?:password|token|apiKey|secret)', stripped, re.I):
                self.add("Medium", "Sensitive Data Exposure",
                         "Sensitive Data in Console Logs", fpath, i, stripped,
                         "Potentially sensitive values printed to browser console",
                         "Remove sensitive data from console output in production builds")

            # ── 4. Input Validation ──────────────────────────────────────
            # JSON.parse without try-catch (look for bare JSON.parse not inside try)
            if "JSON.parse" in stripped and "try" not in stripped:
                # Check if surrounding lines have try
                context = "\n".join(lines[max(0,i-3):i+1])
                if "try" not in context:
                    self.add("Low", "Input Validation",
                             "Unguarded JSON.parse", fpath, i, stripped,
                             "JSON.parse without try-catch can crash on malformed data",
                             "Wrap JSON.parse in try-catch with fallback default")

            # parseInt without radix
            if re.search(r'parseInt\([^,)]+\)', stripped):
                self.add("Low", "Input Validation",
                         "parseInt Without Radix", fpath, i, stripped,
                         "parseInt without explicit radix can misinterpret input",
                         "Always specify radix: parseInt(value, 10)")

            # File upload without type validation
            if re.search(r'accept=.*\*', stripped):
                self.add("Medium", "Input Validation",
                         "Unrestricted File Upload Accept", fpath, i, stripped,
                         "File input accepts all file types without restriction",
                         "Restrict accept attribute to specific MIME types and validate server-side")

            # readAsDataURL without size check
            if "readAsDataURL" in stripped:
                context = "\n".join(lines[max(0,i-5):i])
                if "size" not in context and "maxSize" not in context:
                    self.add("Medium", "Input Validation",
                             "File Upload Without Size Limit", fpath, i, stripped,
                             "File read as DataURL without checking file size first",
                             "Validate file.size < maxAllowed before reading")

            # ── 5. API Security ──────────────────────────────────────────
            # No rate limiting on API calls
            if "generativelanguage.googleapis.com" in stripped:
                self.add("Medium", "API Security",
                         "No Client-Side Rate Limiting", fpath, i, stripped,
                         "External API calls have no rate limiting or debouncing",
                         "Add debounce/throttle to API calls; implement server-side proxy with rate limits")

            # Missing CORS configuration
            if "fetch(" in stripped and "mode" not in stripped:
                if "googleapis" in stripped or "firestore" in stripped:
                    self.add("Low", "API Security",
                             "No Explicit CORS Mode on Fetch", fpath, i, stripped,
                             "fetch() calls without explicit CORS mode rely on browser defaults",
                             "Specify mode: 'cors' and validate server CORS headers")

            # ── 6. Infrastructure & Configuration ────────────────────────
            # Missing .env file
            if "import.meta.env" in stripped:
                env_path = os.path.join(self.root, ".env")
                if not os.path.exists(env_path):
                    self.add("Medium", "Infrastructure",
                             "Missing .env File", fpath, i, stripped,
                             "Code references environment variables but no .env file exists",
                             "Create .env with VITE_FIREBASE_API_KEY and VITE_GEMINI_API_KEY")

            # Missing security headers in HTML
            if fpath.endswith(".html"):
                if "Content-Security-Policy" not in content:
                    if i == 1:  # Only flag once per file
                        self.add("High", "Infrastructure",
                                 "Missing Content-Security-Policy", fpath, i, stripped,
                                 "No CSP meta tag or header to prevent XSS attacks",
                                 "Add <meta http-equiv='Content-Security-Policy' content=\"...\">")

                if "X-Frame-Options" not in content and "frame-ancestors" not in content:
                    if i == 1:
                        self.add("Medium", "Infrastructure",
                                 "Missing X-Frame-Options / Clickjacking Protection", fpath, i, stripped,
                                 "No clickjacking protection headers configured",
                                 "Add X-Frame-Options: DENY or CSP frame-ancestors 'none'")

            # Service worker with overly broad cache
            if "cache.addAll" in stripped:
                self.add("Low", "Infrastructure",
                         "Broad Service Worker Cache", fpath, i, stripped,
                         "Service worker caches root paths which may serve stale content",
                         "Implement cache versioning and network-first strategy for API calls")

            # ── 7. Business Logic Security ───────────────────────────────
            # Patient data in localStorage (PHI exposure)
            if "endopredict_patients" in stripped and "setItem" in stripped:
                self.add("High", "Business Logic",
                         "PHI Stored in localStorage", fpath, i, stripped,
                         "Protected Health Information (patient records) stored in unencrypted localStorage",
                         "Encrypt PHI at rest; use server-side storage with access controls")

            # No audit logging
            if "deleteDoc" in stripped or "deleteAppointment" in stripped:
                self.add("Medium", "Business Logic",
                         "No Audit Trail for Data Deletion", fpath, i, stripped,
                         "Record deletions have no audit log for compliance tracking",
                         "Log all CRUD operations with timestamp, user, and action type")

            # Appointment deletion by index (not ID)
            if "async deleteAppointment(index)" in stripped:
                self.add("Medium", "Business Logic",
                         "Deletion by Array Index (Race Condition)", fpath, i, stripped,
                         "Deleting by array index is fragile; concurrent users may delete wrong record",
                         "Use document ID for deletion instead of positional index")

            # Update by index
            if "async updateAppointment(index" in stripped:
                self.add("Medium", "Business Logic",
                         "Update by Array Index (Race Condition)", fpath, i, stripped,
                         "Updating by array index in concurrent environment risks data corruption",
                         "Use document ID for updates instead of positional index")

            # Client-side trust of user-supplied data
            if "user.password" in stripped and "===" in stripped:
                self.add("Critical", "Business Logic",
                         "Client-Side Password Verification", fpath, i, stripped,
                         "Password verification happens client-side; attacker can bypass",
                         "Move authentication to server-side (Firebase Auth or backend API)")

            # ── 8. Firestore Security ────────────────────────────────────
            # No Firestore security rules reference
            if "getFirestore" in stripped:
                rules_path = os.path.join(self.root, "firestore.rules")
                if not os.path.exists(rules_path):
                    self.add("High", "Infrastructure",
                             "Missing Firestore Security Rules File", fpath, i, stripped,
                             "No firestore.rules file found; database may use default open rules",
                             "Create firestore.rules with proper read/write restrictions per collection")

            # ── 9. Dependency Security ───────────────────────────────────
            if fpath.endswith("package.json") and '"dependencies"' in stripped:
                self.add("Low", "Infrastructure",
                         "Dependency Audit Recommended", fpath, i, stripped,
                         "No evidence of regular npm audit in CI pipeline",
                         "Add 'npm audit --audit-level=high' to CI/CD pipeline")

    def scan_directory(self, dirpath):
        """Recursively scan all source files."""
        for root, dirs, files in os.walk(dirpath):
            # Skip node_modules, dist, .git, venv
            dirs[:] = [d for d in dirs if d not in
                       {"node_modules", "dist", ".git", "venv", "__pycache__", "ios"}]
            for fname in files:
                ext = os.path.splitext(fname)[1]
                if ext in EXTENSIONS:
                    self.scan_file(os.path.join(root, fname))

    def run(self):
        """Execute full scan."""
        print("\n" + "=" * 80)
        print("🔍 ENDOPREDICT STATIC SECURITY CODE REVIEW (SAST)")
        print("=" * 80)
        print(f"  Project Root : {self.root}")
        print(f"  Started At   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 80 + "\n")

        self.scan_directory(self.root)

        # Print findings summary
        by_severity = defaultdict(int)
        for f in self.findings:
            by_severity[f.severity] += 1

        print(f"  Total Findings: {len(self.findings)}")
        for sev in ["Critical", "High", "Medium", "Low"]:
            count = by_severity.get(sev, 0)
            icon = {"Critical": "🔴", "High": "🟠", "Medium": "🟡", "Low": "🔵"}[sev]
            print(f"    {icon} {sev}: {count}")
        print()

        # Print each finding
        for f in self.findings:
            icon = {"Critical": "🔴", "High": "🟠", "Medium": "🟡", "Low": "🔵"}[f.severity]
            print(f"  {icon} [{f.fid}] {f.severity} — {f.vuln_type}")
            print(f"     File: {f.file_path}:{f.line_no}")
            print(f"     {f.explanation}")
            print()

        print("=" * 80)
        return self.findings


# ── Excel Report ──────────────────────────────────────────────────────────────
def generate_excel(findings, project_root):
    wb = openpyxl.Workbook()

    title_font   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font    = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=10)
    navy_fill    = PatternFill("solid", fgColor="1B365D")
    blue_fill    = PatternFill("solid", fgColor="1A73E8")
    grey_fill    = PatternFill("solid", fgColor="F1F3F4")
    crit_fill    = PatternFill("solid", fgColor="F8D7DA")
    high_fill    = PatternFill("solid", fgColor="FFE0CC")
    med_fill     = PatternFill("solid", fgColor="FFF3CD")
    low_fill     = PatternFill("solid", fgColor="D1ECF1")
    crit_font    = Font(name="Calibri", size=10, bold=True, color="721C24")
    high_font    = Font(name="Calibri", size=10, bold=True, color="856404")
    med_font     = Font(name="Calibri", size=10, bold=True, color="856404")
    low_font     = Font(name="Calibri", size=10, bold=True, color="0C5460")
    thin_border  = Border(
        left=Side("thin", "D0D0D0"), right=Side("thin", "D0D0D0"),
        top=Side("thin", "D0D0D0"), bottom=Side("thin", "D0D0D0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sev_style = {
        "Critical": (crit_fill, crit_font),
        "High":     (high_fill, high_font),
        "Medium":   (med_fill, med_font),
        "Low":      (low_fill, low_font),
    }

    by_sev = defaultdict(int)
    by_cat = defaultdict(int)
    for f in findings:
        by_sev[f.severity] += 1
        by_cat[f.category] += 1

    # ── Sheet 1: Executive Summary ───────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws.merge_cells("A1:H2")
    c = ws["A1"]
    c.value = f"  🛡️ EndoPredict Static Security Review — {len(findings)} Findings"
    c.font = title_font; c.fill = navy_fill; c.alignment = Alignment(vertical="center")

    # Severity breakdown
    kpis = [
        ("Total Findings",   len(findings),         "B4", "C4"),
        ("Critical",         by_sev.get("Critical",0), "E4", "F4"),
        ("High",             by_sev.get("High",0),     "B6", "C6"),
        ("Medium",           by_sev.get("Medium",0),   "E6", "F6"),
        ("Low",              by_sev.get("Low",0),      "B8", "C8"),
        ("Scan Date",        datetime.now().strftime("%Y-%m-%d %H:%M"), "E8", "F8"),
    ]
    for label, val, lc, vc in kpis:
        ws[lc].value = label; ws[lc].font = bold_font; ws[lc].fill = grey_fill
        ws[lc].alignment = center; ws[lc].border = thin_border
        ws[vc].value = val; ws[vc].font = bold_font
        ws[vc].alignment = center; ws[vc].border = thin_border

    # Category breakdown table
    row = 11
    ws.cell(row, 1, "Findings by Category").font = Font(size=13, bold=True, color="1B365D")
    row += 1
    for ci, h in enumerate(["Category", "Count", "Critical", "High", "Medium", "Low"], 1):
        c = ws.cell(row, ci, h); c.font = header_font; c.fill = blue_fill
        c.alignment = center; c.border = thin_border
    row += 1
    for cat in sorted(by_cat.keys()):
        cat_findings = [f for f in findings if f.category == cat]
        sev_counts = defaultdict(int)
        for f in cat_findings:
            sev_counts[f.severity] += 1
        vals = [cat, len(cat_findings), sev_counts["Critical"],
                sev_counts["High"], sev_counts["Medium"], sev_counts["Low"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci, v); c.font = regular_font
            c.alignment = center if ci > 1 else left; c.border = thin_border
        row += 1

    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mx + 4, 14)

    # ── Sheet 2: Detailed Findings ───────────────────────────────────────
    ws2 = wb.create_sheet(f"Findings ({len(findings)})")
    headers = ["ID", "Severity", "Category", "Vulnerability Type",
               "File", "Line", "Code Snippet", "Explanation", "Remediation"]
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(1, ci, h); c.font = header_font; c.fill = navy_fill
        c.alignment = center; c.border = thin_border

    for ri, f in enumerate(findings, 2):
        vals = [f.fid, f.severity, f.category, f.vuln_type,
                f.file_path, f.line_no, f.code_snippet,
                f.explanation, f.remediation]
        fill, sfont = sev_style[f.severity]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci, v); c.font = regular_font; c.border = thin_border
            if ci == 2:
                c.fill = fill; c.font = sfont
            c.alignment = center if ci in [1,2,3,6] else left

    for col in ws2.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 4, 12), 55)

    path = os.path.join(project_root, REPORT_NAME)
    wb.save(path)
    return path


# ── Markdown Executive Summary ────────────────────────────────────────────────
def generate_markdown_summary(findings, project_root):
    os.makedirs(SUMMARY_DIR, exist_ok=True)

    by_sev = defaultdict(int)
    for f in findings:
        by_sev[f.severity] += 1

    lines = [
        "# EndoPredict AI — Static Security Code Review",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Total Findings:** {len(findings)}",
        "",
        "## Severity Summary",
        "",
        "| Severity | Count |",
        "|----------|-------|",
        f"| 🔴 Critical | {by_sev.get('Critical', 0)} |",
        f"| 🟠 High | {by_sev.get('High', 0)} |",
        f"| 🟡 Medium | {by_sev.get('Medium', 0)} |",
        f"| 🔵 Low | {by_sev.get('Low', 0)} |",
        "",
        "## Top Findings",
        "",
    ]

    for f in findings:
        if f.severity in ("Critical", "High"):
            icon = "🔴" if f.severity == "Critical" else "🟠"
            lines.append(f"### {icon} [{f.fid}] {f.vuln_type}")
            lines.append(f"- **Severity:** {f.severity}")
            lines.append(f"- **File:** `{f.file_path}:{f.line_no}`")
            lines.append(f"- **Issue:** {f.explanation}")
            lines.append(f"- **Fix:** {f.remediation}")
            lines.append("")

    lines.extend([
        "## Recommendations",
        "",
        "1. **Implement server-side authentication** using Firebase Auth instead of client-side password comparison",
        "2. **Hash all passwords** with bcrypt/argon2 before storage in Firestore",
        "3. **Proxy API calls** through a backend to avoid exposing API keys in the browser",
        "4. **Add Content-Security-Policy** headers to prevent XSS attacks",
        "5. **Encrypt PHI** (Protected Health Information) at rest in localStorage or migrate to server-side storage",
        "6. **Create Firestore security rules** to restrict database access by authenticated user role",
        "7. **Add session expiry** to prevent indefinite authentication persistence",
        "8. **Run `npm audit`** regularly and add it to your CI/CD pipeline",
        "",
        "---",
        f"*Report generated by EndoPredict SAST Scanner on {datetime.now().strftime('%Y-%m-%d')}*",
    ])

    summary_path = os.path.join(SUMMARY_DIR, "Executive_Summary.md")
    with open(summary_path, "w") as f:
        f.write("\n".join(lines))
    return summary_path


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        scanner = SecurityScanner(PROJECT_ROOT)
        findings = scanner.run()

        excel_path = generate_excel(findings, PROJECT_ROOT)
        print(f"\n📊 Excel Report: {excel_path}")

        md_path = generate_markdown_summary(findings, PROJECT_ROOT)
        print(f"📝 Executive Summary: {md_path}")

        print(f"\n✅ Static security review completed with {len(findings)} findings.")
    except Exception as e:
        print(f"\n❌ Security review failed: {e}")
        traceback.print_exc()
        sys.exit(1)
